# -*- coding: utf-8 -*-

"""
Module implementing label.
"""
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QMainWindow,QMessageBox,  QFileDialog
from PyQt5 import QtGui
from .Ui_label import Ui_MainWindow
from openpyxl.styles import Font
from openpyxl import load_workbook
import os
import win32print

class label(QMainWindow, Ui_MainWindow):
    """
    Class documentation goes here.
    """
    def __init__(self, parent=None):
        """
        Constructor
        
        @param parent reference to the parent widget
        @type QWidget
        """
        super(label, self).__init__(parent)
        self.setupUi(self)
        self.name="\\\\Polyzoudis-pc\\Ετικέτες"
    @pyqtSlot()
    def on_bt_browser_clicked(self):
        """
        get the excel and collect the data
        """
        #anixe to arxeio
        self.filepath = QFileDialog.getOpenFileName(self, "ΑΝΟΙΞΕ ΤΟ", "ΑΝΟΙΞΕ ΤΟ", "ΤΑΣΟ (*.xlsx)")[0]
        print(self.filepath)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.data_browser.setFont(font)
        self.data_browser.setText(self.filepath)
       
        #pare ta stoixeia
        wb = load_workbook(filename=self.filepath, read_only=True, data_only=True)
        ws =wb.active
        self.line_1.setText(ws['C2'].value)
        self.line_2.setText (ws['C3'].value)
        self.line_3.setText(ws['C4'].value)
        ask_day = ws['C5'].value.strftime('%m/%d/%Y')
        self.line_4.setText(ask_day)
        delivery_day = ws['C6'].value.strftime('%m/%d/%Y')
        self.line_5.setText(delivery_day)
        self.line_6.setText(ws['C7'].value)
        self.line_7.setText(ws['C8'].value)
        print(os.path.basename(self.filepath))
        self.line_8.setText (os.path.basename(self.filepath))


    @pyqtSlot()
    def on_bt_print_clicked(self):
        """
        Slot documentation goes here.
        """     
        a = "^XA^FX #^CF0,60^FO20,30^FDMILKPLAN^FS^CF0,30^FO620,55^FDRnD_CODES^FS^FO20,85^GB750,1,3^FS^FX #"
        b = "^CFA,28^FO50,120^FD"+"AITHSH KOPHS:"+self.line_1.text()+"^FS"
        c ="^FO50,160^FD"+"ENTOLH: "+self.line_2.text()+"^FS"
        d = "^FO50,200^FD"+"TOMEAS: "+self.line_3.text()+"^FS"
        e = "^FO50,240^FD"+"DELIVERY DATE:"+self.line_5.text()+"^FS"
        e1 = "^FO50,280^FD"+"NAME:"+self.line_6.text()+"^FS"
        e2 = "^FO50,320^FD"+"DELIVERY PLACE:"+self.line_7.text()+"^FS"
        f = "^CFA,25^FO20,450^FDCOMMENTS:^FS^FX Fourth section (the two boxes on the bottom)."
        g = "^FO20,480^GB750,100,5^FS^CF0,40^FO50,510^FD"+self.data_paratirisis.text()+"^FS^XZ"

        message = a+b+c+d+e+e1+e2+f+g
#        hitzebra = win32print.OpenPrinter("\\\\Polyzoudis-pc\\Ετικέτες")
#        hitzebra = win32print.SetDefaultPrinter()
        hitzebra = win32print.OpenPrinter(self.name)
        filename = "2222.txt"
        s = bytes(message, "utf-16")
        print(type(s))
        print(s)
        print(hitzebra)
        posot=int(self.data_posothta.text())
        print(posot) 
        try:
            hJob = win32print.StartDocPrinter(hitzebra, 1, ('PrintJobName', None, 'RAW'))
            try:
#                win32api.ShellExecute(0, "print", filename, None, ".", 0)
             for i in range(posot):
                win32print.StartPagePrinter(hitzebra)
                win32print.WritePrinter(hitzebra, s)  # Instead of raw text is there a way to print PDF File ?
                win32print.EndPagePrinter(hitzebra)
                 
            finally:
                win32print.EndDocPrinter(hitzebra)
        finally:
#            win32print.ClosePrinter("\\\\Polyzoudis-pc\\Ετικέτες")
             pass

    @pyqtSlot()
    def on_bt_green_clicked(self):
        """
        makew the excel paper green
        """
        wb = load_workbook(filename=self.filepath)
        ws =wb.active
        col = ws.column_dimensions['A':'W']
        ft = Font(color=colors.RED)
        col.font = ft
        wb.save(filename = self.filepath)  
    @pyqtSlot()
    def on_bt_about_clicked(self):
        """
        ABOUT ME and selects the zebra printer
        """
        xprinters = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL, None, 1)
        print (xprinters)
        printers = win32print.EnumPrinters(5)
        current = printers[-1]
        name = current[-2]
        self.name = name 
        self.printer = printers
        print (printers)
        print(current)
        print(name)
        
        QMessageBox.about(self, "copyright_A_TMHMATOS", "creator αμανατιδης αλκης \nemail:amanatidisalkis@gmail.com ")
